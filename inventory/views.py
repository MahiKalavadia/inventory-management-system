from django.shortcuts import render, redirect, get_object_or_404
from .models import Product, Category, StockLog
from suppliers.models import Supplier
from django.db.models import Q, F, Sum,  FloatField, ExpressionWrapper, DecimalField, Value, Avg, Max
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from .forms import ProductForm, CategoryForm, StockForm
from django.db.models import Count
from inventory.config import get_low_stock_threshold
from django.contrib import messages
from django.db.models.functions import TruncMonth, Coalesce
from django.http import HttpResponse
import csv
import json
from openpyxl import Workbook
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.utils.timezone import now
from datetime import timedelta
from orders.models import Order
from purchases.models import PurchaseRequest
from django.utils import timezone
from datetime import timedelta


@login_required
def product_dashboard(request):
    # Search & Filters
    search = request.GET.get("search", "")
    category_id = request.GET.get("category", "")
    supplier_id = request.GET.get("supplier", "")
    status = request.GET.get("status", "")
    categories = Category.objects.all()
    suppliers = Supplier.objects.all()

    products = Product.objects.select_related(
        "category", "supplier").filter(is_active=True)

    if search:
        products = products.filter(
            Q(name__icontains=search) |
            Q(brand__icontains=search)
        )

    if category_id:
        products = products.filter(category_id=category_id)

    if supplier_id:
        products = products.filter(supplier_id=supplier_id)

    if status == "in":
        products = products.filter(quantity__gt=get_low_stock_threshold())
    elif status == "low":
        products = products.filter(
            quantity__lte=get_low_stock_threshold(), quantity__gt=0)
    elif status == "out":
        products = products.filter(quantity=0)

    # Pagination
    paginator = Paginator(products, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Stats
    total_products = Product.objects.count()
    in_stock = Product.objects.filter(
        quantity__gt=get_low_stock_threshold(), is_active=True).count()
    low_stock = Product.objects.filter(
        quantity__lte=get_low_stock_threshold(), quantity__gt=0, is_active=True).count()
    out_of_stock = Product.objects.filter(quantity=0, is_active=True).count()
    total_categories = Category.objects.count()
    total_suppliers = Supplier.objects.count()
    suppliers = Supplier.objects.all()

    # charts
    top_categories = (
        Product.objects.values('category__name')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )

    category_names = [c['category__name']
                      or "Uncategorized" for c in top_categories]
    category_counts = [c['total'] for c in top_categories]

    top_expensive = Product.objects.filter(
        price__isnull=False).order_by('-price')[:5]
    expensive_names = [p.name for p in top_expensive]
    expensive_prices = [float(p.price) for p in top_expensive]

    monthly_products = (
        Product.objects
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(total=Count('id'))
        .order_by('month')
    )

    months = [m['month'].strftime('%b %Y')
              for m in monthly_products if m['month']]
    monthly_counts = [m['total'] for m in monthly_products]

    category_value_data = Category.objects.annotate(
        inventory_value=Sum(
            ExpressionWrapper(
                F('product__price') * F('product__quantity'),
                output_field=FloatField()
            )
        )
    ).order_by('-inventory_value')[:5]

    total_inventory_value = Product.objects.aggregate(
        value=Sum(
            ExpressionWrapper(
                F('price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    )['value'] or 0

    context = {
        'products': page_obj,
        'page_obj': page_obj,
        'categories': categories,
        "suppliers": suppliers,
        "total_products": total_products,
        "in_stock": in_stock,
        "low_stock": low_stock,
        "out_of_stock": out_of_stock,
        "total_categories": total_categories,
        "total_suppliers": total_suppliers,
        "low_stock_limit": get_low_stock_threshold(),
        'category_names': category_names,
        'category_counts': category_counts,
        'expensive_names': expensive_names,
        'expensive_prices': expensive_prices,
        'months': months,
        'monthly_counts': monthly_counts,
        'category_value_labels': [c.name for c in category_value_data],
        'category_value_values': [float(c.inventory_value or 0) for c in category_value_data],
        'total_inventory_value': total_inventory_value,
    }

    return render(request, "dashboards/product_dashboard.html", context)


@login_required
def product_list(request):
    products = Product.objects.select_related(
        'category', 'supplier').filter(is_active=True)

    search = request.GET.get('search')
    category = request.GET.get('category')
    supplier = request.GET.get('supplier')
    status = request.GET.get('status')

    if search:
        products = products.filter(name__icontains=search)

    if category:
        products = products.filter(category_id=category)

    if supplier:
        products = products.filter(supplier_id=supplier)

    if status == 'in':
        products = products.filter(quantity__gt=get_low_stock_threshold())
    elif status == 'low':
        products = products.filter(
            quantity__gt=0, quantity__lte=get_low_stock_threshold())
    elif status == 'out':
        products = products.filter(quantity=0)

    paginator = Paginator(products, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory/product_list.html', {
        'page_obj': page_obj,
        "low_stock_limit": get_low_stock_threshold(),
        'categories': Category.objects.all(),
        'suppliers': Supplier.objects.all(),
    })


@login_required
def add_product(request):
    if request.method == "POST":
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('product_dashboard')
    else:
        form = ProductForm()

    return render(request, 'inventory/add_product.html', {'form': form})


@login_required
def update_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            return redirect('product_dashboard')
    else:
        form = ProductForm(instance=product)

    return render(request, 'inventory/update_product.html', {'form': form})


@login_required
def delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':
        product.is_active = False
        product.save()
        return redirect('product_dashboard')

    return render(request, 'inventory/delete_product.html', {'product': product})


@login_required
def top_value_products(request):
    top_value_products = Product.objects.annotate(
        value=ExpressionWrapper(
            F('purchase_price') * F('quantity'),
            output_field=FloatField()
        )
    ).order_by('-value')

    paginator = Paginator(top_value_products, 15)
    page_number = request.GET.get('page')
    productt = paginator.get_page(page_number)

    context = {
        'productt': productt,
    }
    return render(request, 'inventory/top_value_products.html', context)


def export_product_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="products.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "SKU", "Name", "Brand", "Purchase Price", "Selling Price", "Profit", "Quantity", "Category", "Supplier", "Warranty Months"
    ])

    for product in Product.objects.all():
        writer.writerow([
            product.sku, product.name, product.brand, product.purchase_price, product.price, product.profit, product.quantity, product.category, product.supplier, product.warranty_months
        ])

    return response


def export_product_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    ws.append(["SKU", "Name", "Brand", "Purchase Price", "Selling Price",
              "Profit", "Quantity", "Category", "Supplier", "Warranty Months"])

    for product in Product.objects.all():
        ws.append([
            product.sku,
            product.name,
            product.brand,
            product.purchase_price,
            product.price,
            product.profit,
            product.quantity,
            product.category.name,
            product.supplier.name,
            product.warranty_months,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="products.xlsx"'

    wb.save(response)
    return response


def export_product_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="products.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))  # landscape mode
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("PRODUCT LIST", styles["Heading1"]))
    elements.append(Spacer(1, 10))

    # Table Header
    data = [["SKU", "Name", "Brand", "Purchase Price", "Price",
             "Profit", "Qty", "Category", "Supplier", "Warranty (Months)"]]

    products = Product.objects.all()
    for item in products:
        data.append([
            item.sku,
            item.name,
            item.brand,
            item.purchase_price,
            item.price,
            item.profit,
            item.quantity,
            item.category.name if item.category else "",
            item.supplier.name if item.supplier else "",
            item.warranty_months,
        ])

    # Column widths for landscape
    col_widths = [60, 150, 80, 80, 80, 80, 50, 100, 100, 60]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([

        # Header Style
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F3E46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        # Body Style
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),

        # Row alternating light grey
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F7F9FB")]),

        # Align price column right
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),

        # Subtle grid
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#B0BEC5")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),

        # Padding
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

    ]))

    elements.append(table)
    doc.build(elements)
    return response


@login_required
def category_dashboard(request):
    categories_qs = Category.objects.annotate(
        product_count=Count('product', distinct=True)
    ).order_by('name')

    search = request.GET.get('q')
    if search:
        categories_qs = categories_qs.filter(
            Q(name__icontains=search)
        )

    paginator = Paginator(categories_qs, 5)
    page_number = request.GET.get('page')
    categoriess = paginator.get_page(page_number)

    total_category = Category.objects.count()
    categories_with_products = Category.objects.annotate(
        product_count=Count('product')
    ).filter(product_count__gt=0).count()
    categories_without_products = Category.objects.annotate(
        product_count=Count('product')
    ).filter(product_count=0).count()

    # Charts
    # stock per category
    categories = list(
        Category.objects.annotate(
            total_stock=Coalesce(Sum('product__quantity'), 0),
            total_stock_value=Coalesce(
                Sum(
                    F('product__price') * F('product__quantity'),
                    output_field=DecimalField(max_digits=12, decimal_places=2)
                ),
                Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=2)
            ),
            total_sales=Coalesce(
                Sum(
                    F('product__orderitem__price') *
                    F('product__orderitem__quantity'),
                    output_field=DecimalField(max_digits=12, decimal_places=2)
                ),
                Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        ).order_by('-total_sales')[:5].values(
            'name',
            'total_stock',
            'total_stock_value',
            'total_sales'
        )
    )

    category_names = [c['name'] for c in categories]
    stock_totals = [c['total_stock'] for c in categories]
    stock_values = [float(c['total_stock_value']) for c in categories]
    sales_values = [float(c['total_sales']) for c in categories]

    context = {
        'categoriess': categoriess,
        'total_category': total_category,
        'categories_with_products': categories_with_products,
        'categories_without_products': categories_without_products,
        'search': search,
        'category_names': json.dumps(category_names),
        'stock_totals': json.dumps(stock_totals),
        'stock_values': json.dumps(stock_values),
        'sales_values': json.dumps(sales_values),
    }

    return render(request, "dashboards/category_dashboard.html", context)


@login_required
def category_without_products(request):
    categories_qs = Category.objects.annotate(
        product_count=Count('product', distinct=True)
    ).order_by('name')
    categories_without_products = categories_qs.filter(product_count=0).all()

    paginator = Paginator(categories_without_products, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "inventory/category_without_products.html", {'page_obj': page_obj})


@login_required
def category_with_products(request):
    categories_qs = Category.objects.annotate(
        product_count=Count('product', distinct=True)
    ).order_by('name')
    categories_with_products = categories_qs.filter(
        product_count__gt=0).all()

    paginator = Paginator(categories_with_products, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "inventory/category_with_products.html", {'page_obj': page_obj})


@login_required
def category_list(request):
    categories = Category.objects.all()
    data = []

    for category in categories:
        product_count = Product.objects.filter(category=category).count()
        data.append({
            'category': category,
            'total_products': product_count
        })
    paginator = Paginator(data, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory/category_list.html', {
        'page_obj': page_obj,
        'data': data
    })


@login_required
def category_products(request, category_id):
    category = get_object_or_404(Category, id=category_id)

    product_list = Product.objects.filter(category=category, is_active=True)

    paginator = Paginator(product_list, 10)
    page_number = request.GET.get('page')
    products = paginator.get_page(page_number)

    return render(request, 'inventory/category_products.html', {
        'category': category,
        'products': products
    })


@login_required
def add_category(request):
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('category_dashboard')
    else:
        form = CategoryForm()

    return render(request, 'inventory/add_category.html', {'form': form})


@login_required
def update_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            form.save()
            return redirect('category_dashboard')
    else:
        form = CategoryForm(instance=category)

    return render(request, 'inventory/update_category.html', {'form': form})


@login_required
def delete_category(request, pk):
    category = get_object_or_404(Category, pk=pk)

    if request.method == 'POST':
        category.delete()
        return redirect('category_dashboard')

    return render(request, 'inventory/delete_category.html', {
        'category': category
    })


def export_category_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="category.csv"'

    writer = csv.writer(response)
    writer.writerow(["Name"])

    for cate in Category.objects.all():
        writer.writerow([
            cate.name,
        ])

    return response


def export_category_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Categories"

    ws.append(["Name"])

    for category in Category.objects.all():
        ws.append([category.name])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="categories.xlsx"'

    wb.save(response)
    return response


def export_category_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="category.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))  # landscape mode
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("CATEGORY LIST", styles["Heading1"]))
    elements.append(Spacer(1, 10))

    # Table Header
    data = [["Name"]]

    category = Category.objects.all()
    for cat in category:
        data.append([
            cat.name,
        ])

    # Column widths for landscape
    col_widths = [300]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([

        # Header Style
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F3E46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        # Body Style
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),

        # Row alternating light grey
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F7F9FB")]),

        # Align price column right
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),

        # Subtle grid
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#B0BEC5")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),

        # Padding
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

    ]))

    elements.append(table)
    doc.build(elements)
    return response


@login_required
def stock_dashboard(request):
    # ---------- FILTER PARAMETERS ----------
    category_id = request.GET.get('category')
    supplier_id = request.GET.get('supplier')
    status = request.GET.get('status')
    search = request.GET.get('q')

    # ---------- BASE QUERY ----------
    product_qs = Product.objects.select_related(
        'category', 'supplier').filter(is_active=True)

    # ---------- APPLY FILTERS ----------
    if category_id:
        product_qs = product_qs.filter(category_id=category_id)
    if supplier_id:
        product_qs = product_qs.filter(supplier_id=supplier_id)
    if status == 'in':
        product_qs = product_qs.filter(quantity__gt=get_low_stock_threshold())
    elif status == 'low':
        product_qs = product_qs.filter(
            quantity__lte=get_low_stock_threshold(), quantity__gt=0)
    elif status == 'out':
        product_qs = product_qs.filter(quantity=0)
    if search:
        product_qs = product_qs.filter(
            Q(name__icontains=search) | Q(sku__icontains=search))

    # ---------- SUMMARY COUNTS ----------
    total_products = Product.objects.filter(is_active=True).count()
    in_stock = Product.objects.filter(
        quantity__gt=get_low_stock_threshold(), is_active=True).count()
    low_stock = Product.objects.filter(
        quantity__lte=get_low_stock_threshold(), quantity__gt=0, is_active=True).count()
    out_stock = Product.objects.filter(quantity=0, is_active=True).count()

    # ---------- TOTAL STOCK IN/OUT ----------
    total_stock_in = StockLog.objects.filter(action='IN').aggregate(
        Sum('quantity'))['quantity__sum'] or 0
    total_stock_out = StockLog.objects.filter(
        action='OUT').aggregate(Sum('quantity'))['quantity__sum'] or 0

    # ---------- LOW STOCK PRODUCTS ----------
    low_qs = Product.objects.filter(
        quantity__lte=get_low_stock_threshold(), quantity__gt=0, is_active=True).select_related('category', 'supplier')

    # ---------- PAGINATION ----------
    product_paginator = Paginator(product_qs.order_by('name'), 5)
    product_page = request.GET.get('page_products')
    products = product_paginator.get_page(product_page)

    low_paginator = Paginator(low_qs.order_by('name'), 5)
    low_page = request.GET.get('page_low')
    low_stock_products = low_paginator.get_page(low_page)

    # ---------- RECENT LOGS ----------
    recent_logs = StockLog.objects.select_related(
        'product', 'user').order_by('-created_at')[:10]

    # Charts
    category_data = (
        Product.objects
        .values('category__name')
        .annotate(
            total_value=Sum(
                ExpressionWrapper(
                    F('quantity') * F('price'),
                    output_field=DecimalField()
                )
            )
        )
        .order_by('-total_value')[:5]
    )

    category_labels = []
    category_values = []

    for item in category_data:
        category_labels.append(item['category__name'] or "No Category")
        category_values.append(float(item['total_value'] or 0))

    # ============================
    # 3️⃣ TOP 5 LOW STOCK PRODUCTS
    # ============================

    low_products = (
        Product.objects
        .filter(quantity__gt=0)
        .order_by('quantity')[:5]
    )

    low_product_labels = [p.name for p in low_products]
    low_product_values = [p.quantity for p in low_products]

    # ============================
    # 4️⃣ MONTHLY STOCK SUPPLIED
    # ============================
    last_six_months = timezone.now() - timedelta(days=180)
    monthly_data = (
        PurchaseRequest.objects
        .filter(status="Approved", created_at__gte=last_six_months)
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(total=Sum('quantity'))
        .order_by('month')
    )

    monthly_labels = []
    monthly_supplied = []

    for item in monthly_data:
        month_name = item['month'].strftime('%b %Y')
        monthly_labels.append(month_name)
        monthly_supplied.append(item['total'] or 0)

    # ---------- CONTEXT ----------
    context = {
        # cards
        'total_products': total_products,
        'in_stock': in_stock,
        'low_stock': low_stock,
        'out_stock': out_stock,

        # tables
        'products': products,
        'low_stock_products': low_stock_products,
        'recent_logs': recent_logs,

        # stats
        'total_stock_in': total_stock_in,
        'total_stock_out': total_stock_out,

        # filters
        'categories': Category.objects.all(),
        'suppliers': Supplier.objects.all(),

        "category_labels": category_labels,
        "category_values": category_values,

        "low_product_labels": low_product_labels,
        "low_product_values": low_product_values,

        "monthly_labels": monthly_labels,
        "monthly_supplied": monthly_supplied,
    }

    return render(request, 'dashboards/stock_dashboard.html', context)


@login_required
def stock_in(request, product_id=None):
    product_instance = None
    if product_id:
        product_instance = get_object_or_404(Product, id=product_id)

    form = StockForm(request.POST)
    if request.method == 'POST' and form.is_valid():
        product = form.cleaned_data['product'] if not product_instance else product_instance
        qty = form.cleaned_data['quantity']

        product.quantity = F('quantity') + qty
        product.save(update_fields=['quantity'])

        StockLog.objects.create(
            product=product,
            action=StockLog.STOCK_IN,
            quantity=qty
        )

        messages.success(request, "Stock added successfully")
        user = request.user

        if user.is_superuser:
            return redirect('admin_dashboard')
        elif user.is_staff:
            return redirect('manager_dashboard')
        else:
            return redirect('staff_dashboard')
    else:
        # Pre-fill form with product if available
        initial_data = {
            'product': product_instance} if product_instance else {}
        form = StockForm(initial=initial_data)
    return render(request, 'inventory/stock_in.html', {'form': form, 'product': product_id})


@login_required
def stock_out(request):
    form = StockForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        product = form.cleaned_data['product']
        qty = form.cleaned_data['quantity']

        product.refresh_from_db()

        if qty > product.quantity:
            messages.error(request, 'Insufficient stock availability!')
            return redirect('stock_out')

        product.quantity = F('quantity') - qty
        product.save(update_fields=['quantity'])

        StockLog.objects.create(
            product=product,
            action=StockLog.STOCK_OUT,
            quantity=qty
        )

        messages.success(request, "Stock removed successfully!")

        # role based redirect
        user = request.user
        if user.is_superuser:
            return redirect('admin_dashboard')
        elif user.is_staff:
            return redirect('manager_dashboard')
        else:
            return redirect('staff_dashboard')

    return render(request, 'inventory/stock_out.html', {'form': form})


@login_required
def stock_history(request):
    logs = StockLog.objects.select_related(
        'product', 'user').order_by('-created_at')

    paginator = Paginator(logs, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory/stock_history.html', {'page_obj': page_obj})

@login_required
def in_stock_products(request):
    products = Product.objects.filter(
        quantity__gt=get_low_stock_threshold(), is_active=True)

    MAX_STOCK = 100

    for p in products:
        p.stock_percent = (p.quantity / MAX_STOCK) * 100

    paginator = Paginator(products, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory/in_stock.html', {'page_obj': page_obj})


@login_required
def low_stock_products(request):
    products = Product.objects.filter(
        quantity__lte=get_low_stock_threshold(), quantity__gt=0, is_active=True)

    MAX_STOCK = 10

    for p in products:
        p.stock_percent = (p.quantity / MAX_STOCK) * 100

    paginator = Paginator(products, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory/low_stock.html', {'page_obj': page_obj, "low_stock_limit": get_low_stock_threshold()})


@login_required
def out_stock_products(request):
    products = Product.objects.filter(quantity=0, is_active=True)

    paginator = Paginator(products, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "inventory/out_stock.html", {"page_obj": page_obj})


def export_instock_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filname="instock.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "SKU", "Name", "Brand", "Purchase Price", "Selling Price",
        "Profit", "Quantity", "Category", "Supplier", "Warranty Months"
    ])

    for instock in Product.objects.filter(quantity__gt=get_low_stock_threshold()).all():
        writer.writerow([
            instock.sku,
            instock.name,
            instock.brand,
            instock.purchase_price,
            instock.price,
            instock.profit,
            instock.quantity,
            instock.category,
            instock.supplier,
            instock.warranty_months,
        ])

    return response


def export_instock_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Instock"

    ws.append(["SKU", "Name", "Brand", "Purchase Price", "Selling Price",
               "Profit", "Quantity", "Category", "Supplier", "Warranty Months"
               ])

    for instock in Product.objects.filter(quantity__gt=get_low_stock_threshold()).all():
        ws.append([
            instock.sku,
            instock.name,
            instock.brand,
            instock.purchase_price,
            instock.price,
            instock.profit,
            instock.quantity,
            instock.category.name,
            instock.supplier.name,
            instock.warranty_months,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="instock.xlsx"'

    wb.save(response)
    return response


def export_instock_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="products-instock.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))  # landscape mode
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("PRODUCTS IN STOCK LIST", styles["Heading1"]))
    elements.append(Spacer(1, 10))

    # Table Header
    data = [["SKU", "Name", "Brand", "Purchase Price", "Price",
             "Profit", "Qty", "Category", "Supplier", "Warranty (Months)"]]

    products = Product.objects.filter(
        quantity__gt=get_low_stock_threshold()).all()
    for item in products:
        data.append([
            item.sku,
            item.name,
            item.brand,
            item.purchase_price,
            item.price,
            item.profit,
            item.quantity,
            item.category.name if item.category else "",
            item.supplier.name if item.supplier else "",
            item.warranty_months,
        ])

    # Column widths for landscape
    col_widths = [60, 150, 80, 80, 80, 80, 50, 100, 100, 60]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([

        # Header Style
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F3E46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        # Body Style
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),

        # Row alternating light grey
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F7F9FB")]),

        # Align price column right
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),

        # Subtle grid
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#B0BEC5")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),

        # Padding
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

    ]))

    elements.append(table)
    doc.build(elements)
    return response


def export_lowstock_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="lowstock.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "SKU", "Name", "Brand", "Purchase Price", "Selling Price",
               "Profit", "Quantity", "Category", "Supplier", "Warranty Months"
    ])

    for lowstock in Product.objects.filter(quantity__gt=0, quantity__lte=get_low_stock_threshold(),
                                           ).all():
        writer.writerow([
            lowstock.sku,
            lowstock.name,
            lowstock.brand,
            lowstock.purchase_price,
            lowstock.price,
            lowstock.profit,
            lowstock.quantity,
            lowstock.category,
            lowstock.supplier,
            lowstock.warranty_months,
        ])

    return response


def export_lowstock_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "LowStock"

    ws.append([
        "SKU", "Name", "Brand", "Purchase Price", "Selling Price",
               "Profit", "Quantity", "Category", "Supplier", "Warranty Months"
    ])

    for lowstock in Product.objects.filter(quantity__gt=0, quantity__lte=get_low_stock_threshold()).all():
        ws.append([
            lowstock.sku,
            lowstock.name,
            lowstock.brand,
            lowstock.purchase_price,
            lowstock.price,
            lowstock.profit,
            lowstock.quantity,
            lowstock.category.name,
            lowstock.supplier.name,
            lowstock.warranty_months,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="lowstock.xlsx"'

    wb.save(response)
    return response


def export_lowstock_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="products-lowstock.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))  # landscape mode
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(
        Paragraph("PRODUCTS LOW IN STOCK LIST", styles["Heading1"]))
    elements.append(Spacer(1, 10))

    # Table Header
    data = [["SKU", "Name", "Brand", "Purchase Price", "Price",
             "Profit", "Qty", "Category", "Supplier", "Warranty (Months)"]]

    products = Product.objects.filter(
        quantity__gt=0, quantity__lte=get_low_stock_threshold()).all()
    for item in products:
        data.append([
            item.sku,
            item.name,
            item.brand,
            item.purchase_price,
            item.price,
            item.profit,
            item.quantity,
            item.category.name if item.category else "",
            item.supplier.name if item.supplier else "",
            item.warranty_months,
        ])

    # Column widths for landscape
    col_widths = [60, 150, 80, 80, 80, 80, 50, 100, 100, 60]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([

        # Header Style
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F3E46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        # Body Style
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),

        # Row alternating light grey
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F7F9FB")]),

        # Align price column right
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),

        # Subtle grid
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#B0BEC5")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),

        # Padding
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

    ]))

    elements.append(table)
    doc.build(elements)
    return response


def export_outstock_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="outstock.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "SKU", "Name", "Brand", "Purchase Price", "Selling Price",
               "Profit", "Quantity", "Category", "Supplier", "Warranty Months"
    ])

    for outstock in Product.objects.filter(quantity=0
                                           ).all():
        writer.writerow([
            outstock.sku,
            outstock.name,
            outstock.brand,
            outstock.purchase_price,
            outstock.price,
            outstock.profit,
            outstock.quantity,
            outstock.category,
            outstock.supplier,
            outstock.warranty_months,
        ])

    return response


def export_outstock_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "outStock"

    ws.append([
        "SKU", "Name", "Brand", "Purchase Price", "Selling Price",
               "Profit", "Quantity", "Category", "Supplier", "Warranty Months"
    ])

    for outstock in Product.objects.filter(quantity=0).all():
        ws.append([
            outstock.sku,
            outstock.name,
            outstock.brand,
            outstock.purchase_price,
            outstock.price,
            outstock.profit,
            outstock.quantity,
            outstock.category.name,
            outstock.supplier.name,
            outstock.warranty_months,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="outstock.xlsx"'

    wb.save(response)
    return response


def export_outstock_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="products-outstock.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))  # landscape mode
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(
        Paragraph("PRODUCTS OUT OF STOCK LIST", styles["Heading1"]))
    elements.append(Spacer(1, 10))

    # Table Header
    data = [["SKU", "Name", "Brand", "Purchase Price", "Price",
             "Profit", "Qty", "Category", "Supplier", "Warranty (Months)"]]

    products = Product.objects.filter(quantity=0).all()
    for item in products:
        data.append([
            item.sku,
            item.name,
            item.brand,
            item.purchase_price,
            item.price,
            item.profit,
            item.quantity,
            item.category.name if item.category else "",
            item.supplier.name if item.supplier else "",
            item.warranty_months,
        ])

    # Column widths for landscape
    col_widths = [60, 150, 80, 80, 80, 80, 50, 100, 100, 60]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([

        # Header Style
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F3E46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        # Body Style
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),

        # Row alternating light grey
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F7F9FB")]),

        # Align price column right
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),

        # Subtle grid
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#B0BEC5")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),

        # Padding
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

    ]))

    elements.append(table)
    doc.build(elements)
    return response


@login_required
def report_dashboard(request):
    total_products = Product.objects.count()
    total_orders = Order.objects.count()

    in_stock_products = Product.objects.filter(
        quantity__gt=get_low_stock_threshold(),
        is_active=True
    )

    in_stock = in_stock_products.count()

    avg_price = Product.objects.aggregate(
        average_price=Avg('purchase_price')
    )['average_price'] or 0

    reorder_products = Product.objects.filter(
        quantity__lte=get_low_stock_threshold(),
        is_active=True
    )

    reorder_count = reorder_products.count()

    reorder_cost = reorder_products.aggregate(
        total_reorder=Sum(
            ExpressionWrapper(
                (get_low_stock_threshold() - F('quantity')) * F('purchase_price'),
                output_field=DecimalField()
            )
        )
    )['total_reorder'] or 0

    total_inventory_value = Product.objects.aggregate(
        value=Sum(
            ExpressionWrapper(
                F('price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    )['value'] or 0

    total_inventory_count = Product.objects.aggregate(
        total=Sum('quantity')
    )['total'] or 0

    def percent(count):
        return round((count / total_products) * 100, 1) if total_products else 0

    context = {
        'total_products': total_products,
        'total_orders': total_orders,
        'in_stock': in_stock,
        'in_stock_bar': percent(in_stock),
        'avg_price': avg_price,
        'reorder_count': reorder_count,
        'reorder_cost': reorder_cost,
        'total_inventory_value': total_inventory_value,
        'total_inventory_count': total_inventory_count,
    }

    return render(request, "dashboards/report_dashboard.html", context)


@login_required
def stock_report(request):
    products = Product.objects.filter(is_active=True)

    return render(request, "inventory/stock_report.html", {
        "products": products
    })


@login_required
def export_stock_report_csv(request):
    products = Product.objects.filter(is_active=True)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="stock_report.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "SKU", "Name", "Brand", "Quantity", "Price", "Stock Value"
    ])

    for product in products:
        writer.writerow([
            product.sku,
            product.name,
            product.brand,
            product.quantity,
            product.price,
            product.stock_value,
        ])

    return response


@login_required
def export_stock_report_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    products = Product.objects.filter(is_active=True)

    ws.append(["SKU", "Name", "Brand", "Quantity", "Price", "Stock Value"])
    for product in products:
        ws.append([
            product.sku,
            product.name,
            product.brand,
            product.quantity,
            product.price,
            product.stock_value,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment;filename="stock_report.xlsx"'

    wb.save(response)
    return response


@login_required
def export_stock_report_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="stock_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))  # landscape mode
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(
        Paragraph("STOCK REPORT", styles["Heading1"]))
    elements.append(Spacer(1, 10))

    # Table Header
    data = [["SKU", "Name", "Brand", "Quantity", "Price",
             "Stock Value"]]

    products = Product.objects.filter(is_active=True)
    for item in products:
        data.append([
            item.sku,
            item.name,
            item.brand,
            item.quantity,
            item.price,
            item.stock_value,
        ])

    # Column widths for landscape
    col_widths = [60, 150, 120, 80, 80, 80]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([

        # Header Style
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F3E46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        # Body Style
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),

        # Row alternating light grey
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F7F9FB")]),

        # Align price column right
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),

        # Subtle grid
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#B0BEC5")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),

        # Padding
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

    ]))

    elements.append(table)
    doc.build(elements)
    return response


@login_required
def reorder_report(request):
    products = Product.objects.filter(
        quantity__lte=get_low_stock_threshold(),
        is_active=True
    )

    return render(request, "inventory/reorder_report.html", {
        "products": products,
        "threshold": get_low_stock_threshold()
    })


@login_required
def export_reorder_csv(request):
    products = Product.objects.filter(
        quantity__lte=get_low_stock_threshold(),
        is_active=True
    )

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="reorder_report.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "SKU",
        "Product Name",
        "Quantity",
        "Purchase Price"
    ])

    for product in products:
        writer.writerow([
            product.sku,
            product.name,
            product.quantity,
            product.purchase_price
        ])

    return response


@login_required
def export_reorder_report_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reorder Report"

    products = Product.objects.filter(
        quantity__lte=get_low_stock_threshold(),
        is_active=True
    )

    ws.append([
        "SKU", "Name", "Quantity", "Purchase Price"
    ])

    for product in products:
        ws.append([
            product.sku,
            product.name,
            product.quantity,
            product.purchase_price,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment;filename="reorder_report.xlsx"'

    wb.save(response)
    return response


@login_required
def export_reorder_report_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="reorder_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))  # landscape mode
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(
        Paragraph("REORDER REPORT", styles["Heading1"]))
    elements.append(Spacer(1, 10))

    # Table Header
    data = [["SKU", "Name",  "Quantity", "Purchase Price",
             ]]

    products = Product.objects.filter(
        quantity__lte=get_low_stock_threshold(),
        is_active=True
    )
    for item in products:
        data.append([
            item.sku,
            item.name,
            item.quantity,
            item.purchase_price,
        ])

    # Column widths for landscape
    col_widths = [80, 150, 60, 80]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([

        # Header Style
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F3E46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        # Body Style
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),

        # Row alternating light grey
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F7F9FB")]),

        # Align price column right
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),

        # Subtle grid
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#B0BEC5")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),

        # Padding
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

    ]))

    elements.append(table)
    doc.build(elements)
    return response


@login_required
def inventory_value_report(request):
    products = Product.objects.annotate(
        total_value=ExpressionWrapper(
            F('price') * F('quantity'),
            output_field=DecimalField()
        )
    )

    total_value = products.aggregate(
        total=Sum(
            ExpressionWrapper(
                F('price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    )['total'] or 0

    return render(request, "inventory/inventory_value_report.html", {
        "products": products,
        "total_value": total_value
    })


@login_required
def export_inventory_report_csv(request):
    products = Product.objects.annotate(
        total_value=ExpressionWrapper(
            F('price') * F('quantity'),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
    )

    total_value = products.aggregate(
        total=Sum('total_value')
    )['total'] or 0

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Product", "Quantity", "Price", "Total Value"
    ])

    for inventory in products:
        writer.writerow([
            inventory.name,
            inventory.quantity,
            inventory.price,
            inventory.total_value,
        ])

    writer.writerow([])
    writer.writerow(["Grand Total", "", "", total_value])

    return response


@login_required
def export_inventory_report_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    ws.append(["Name", "Quantity", "Price", "Total Value"])
    products = Product.objects.annotate(
        total_value=ExpressionWrapper(
            F('price') * F('quantity'),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
    )

    total_value = products.aggregate(
        total=Sum('total_value')
    )['total'] or 0

    for inventory in products:
        ws.append([
            inventory.name,
            inventory.quantity,
            inventory.price,
            inventory.total_value,
        ])

    ws.append([])
    ws.append(["Grand Total", "", "", total_value])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment;filename="inventory_report.xlsx"'

    wb.save(response)
    return response


@login_required
def export_inventory_report_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="inventory_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))  # landscape mode
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(
        Paragraph("INVENTORY REPORT", styles["Heading1"]))
    elements.append(Spacer(1, 10))

    # Table Header
    data = [["Name",  "Quantity", "Price", "Total Value"
             ]]

    products = Product.objects.annotate(
        total_value=ExpressionWrapper(
            F('price') * F('quantity'),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
    )

    total_value = products.aggregate(
        total=Sum('total_value')
    )['total'] or 0

    for item in products:
        data.append([
            item.name,
            item.quantity,
            item.price,
            item.total_value,
        ])

        data.append([])
        data.append(["Total Inventory Value", "", "", total_value])

    # Column widths for landscape
    col_widths = [150, 60, 100, 100]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([

        # Header Style
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F3E46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        # Body Style
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),

        # Row alternating light grey
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F7F9FB")]),

        # Align price column right
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),

        # Subtle grid
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#B0BEC5")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),

        # Padding
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

    ]))

    elements.append(table)
    doc.build(elements)
    return response


@login_required
def average_price_report(request):
    products = Product.objects.filter(is_active=True)

    avg_price = products.aggregate(
        average_price=Avg('purchase_price')
    )['average_price'] or 0

    return render(request, "inventory/average_price_report.html", {
        "products": products,
        "avg_price": avg_price
    })


@login_required
def export_averageprice_report_csv(request):
    products = Product.objects.filter(is_active=True)

    avg_price = products.aggregate(
        average_price=Avg('purchase_price')
    )['average_price'] or 0

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="average_price.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Name", "Purchase Price"
    ])

    for average in products:
        writer.writerow([
            average.name,
            average.purchase_price,
        ])

    writer.writerow([])
    writer.writerow(["Total Average Price", avg_price])

    return response


@login_required
def export_averageprice_report_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Average Price"

    ws.append(["Name", "Purchase Price"])
    products = Product.objects.filter(is_active=True)

    avg_price = products.aggregate(
        average_price=Avg('purchase_price')
    )['average_price'] or 0

    for avg in products:
        ws.append([
            avg.name,
            avg.purchase_price,
        ])

        ws.append([])
        ws.append(["Average Price", avg_price])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment;filename="averageprice_report.xlsx"'

    wb.save(response)
    return response


@login_required
def export_averageprice_report_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="average_price.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))  # landscape mode
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(
        Paragraph("INVENTORY REPORT", styles["Heading1"]))
    elements.append(Spacer(1, 10))

    # Table Header
    data = [["Name", "Purchase Price"
             ]]

    products = Product.objects.filter(is_active=True)

    avg_price = products.aggregate(
        average_price=Avg('purchase_price')
    )['average_price'] or 0

    for item in products:
        data.append([
            item.name,
            item.purchase_price
        ])

        data.append([])
        data.append(["Average Price", avg_price])

    # Column widths for landscape
    col_widths = [150, 120]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([

        # Header Style
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F3E46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        # Body Style
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),

        # Row alternating light grey
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F7F9FB")]),

        # Align price column right
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),

        # Subtle grid
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#B0BEC5")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),

        # Padding
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

    ]))

    elements.append(table)
    doc.build(elements)
    return response


@login_required
def dead_stock_report(request):
    sixty_days_ago = now() - timedelta(days=60)

    recent_products = StockLog.objects.filter(
        created_at__gte=sixty_days_ago
    ).values_list('product_id', flat=True)

    products = Product.objects.exclude(
        id__in=recent_products
    ).filter(is_active=True)

    return render(request, "inventory/dead_stock_report.html", {
        "products": products,
        "days": 60
    })


@login_required
def export_deadstock_report_csv(request):
    sixty_days_ago = now() - timedelta(days=60)

    recent_products = StockLog.objects.filter(
        created_at__gte=sixty_days_ago
    ).values_list('product_id', flat=True)

    products = Product.objects.exclude(
        id__in=recent_products
    ).filter(is_active=True)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment;filename="deadstock_report.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "SKU", "Name", "Quantity", "Stock Value"
    ])

    for dead in products:
        writer.writerow([
            dead.sku,
            dead.name,
            dead.quantity,
            dead.stock_value,
        ])

    return response


@login_required
def export_deadstock_report_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dead Stock"

    ws.append(["SKU", "Name", "Quantity", "Stock Value"])
    sixty_days_ago = now() - timedelta(days=60)

    recent_products = StockLog.objects.filter(
        created_at__gte=sixty_days_ago
    ).values_list('product_id', flat=True)

    products = Product.objects.exclude(
        id__in=recent_products
    ).filter(is_active=True)

    for dead in products:
        ws.append([
            dead.sku,
            dead.name,
            dead.quantity,
            dead.stock_value
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment;filename="deadstock_report.xlsx'

    wb.save(response)
    return response


@login_required
def export_deadstock_report_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename="deadstock_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))  # landscape mode
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(
        Paragraph("INVENTORY REPORT", styles["Heading1"]))
    elements.append(Spacer(1, 10))

    # Table Header
    data = [["SKU", "Name", "Quantity" "Stock Value"
             ]]

    sixty_days_ago = now() - timedelta(days=60)

    recent_products = StockLog.objects.filter(
        created_at__gte=sixty_days_ago
    ).values_list('product_id', flat=True)

    products = Product.objects.exclude(
        id__in=recent_products
    ).filter(is_active=True)

    for item in products:
        data.append([
            item.sku,
            item.name,
            item.quantity,
            item.stock_value,
        ])

    # Column widths for landscape
    col_widths = [100, 150, 80, 100]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([

        # Header Style
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F3E46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        # Body Style
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),

        # Row alternating light grey
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F7F9FB")]),

        # Align price column right
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),

        # Subtle grid
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#B0BEC5")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),

        # Padding
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

    ]))

    elements.append(table)
    doc.build(elements)
    return response


@login_required
def view_all_reports(request):
    total_products = Product.objects.count()
    total_orders = Order.objects.count()

    avg_price = Product.objects.aggregate(
        average_price=Avg('purchase_price')
    )['average_price'] or 0

    reorder_products = Product.objects.filter(
        quantity__lte=get_low_stock_threshold(),
        is_active=True
    )

    reorder_count = reorder_products.count()

    reorder_cost = reorder_products.aggregate(
        total_reorder=Sum(
            ExpressionWrapper(
                (get_low_stock_threshold() - F('quantity')) * F('purchase_price'),
                output_field=DecimalField()
            )
        )
    )['total_reorder'] or 0

    total_inventory_value = Product.objects.aggregate(
        value=Sum(
            ExpressionWrapper(
                F('price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    )['value'] or 0

    total_inventory_count = Product.objects.aggregate(
        total=Sum('quantity')
    )['total'] or 0

    context = {
        'total_products': total_products,
        'total_orders': total_orders,
        'avg_price': avg_price,
        'reorder_count': reorder_count,
        'reorder_cost': reorder_cost,
        'total_inventory_value': total_inventory_value,
        'total_inventory_count': total_inventory_count,
    }
    return render(request, 'inventory/view_all_reports.html', context)


@login_required
def stock_forecast(request):
    from orders.models import OrderItem
    
    thirty_days_ago = timezone.now() - timedelta(days=30)
    
    # Calculate sales velocity for each product
    sales_data = OrderItem.objects.filter(
        order__created_at__gte=thirty_days_ago,
        order__payment_status='Paid'
    ).values(
        'product__id',
        'product__name',
        'product__sku',
        'product__quantity',
        'product__category__name',
        'product__image'
    ).annotate(
        total_sold=Sum('quantity')
    ).order_by('-total_sold')
    
    forecast_list = []
    for item in sales_data:
        if item['product__id']:
            daily_avg = item['total_sold'] / 30
            days_until_stockout = item['product__quantity'] / daily_avg if daily_avg > 0 else 999
            forecast_30_days = daily_avg * 30
            
            # Calculate reorder quantity (forecast - current stock, minimum 0)
            reorder_qty = max(0, int(forecast_30_days - item['product__quantity']))
            
            status = 'critical' if days_until_stockout < 7 else 'warning' if days_until_stockout < 14 else 'good'
            
            forecast_list.append({
                'product_id': item['product__id'],
                'name': item['product__name'],
                'sku': item['product__sku'],
                'category': item['product__category__name'],
                'current_stock': item['product__quantity'],
                'sold_30_days': item['total_sold'],
                'daily_avg': round(daily_avg, 2),
                'forecast_30_days': round(forecast_30_days, 0),
                'reorder_qty': reorder_qty,
                'days_until_stockout': round(days_until_stockout, 1),
                'status': status,
                'image': item['product__image']
            })
    
    # Top 10 by highest sales (already sorted by total_sold descending)
    top_10_critical = forecast_list[:10]
    
    # All products sorted by sales velocity
    paginator = Paginator(forecast_list, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Chart data - Top 10 by sales
    chart_labels = [p['name'][:20] for p in top_10_critical]
    chart_sold = [p['sold_30_days'] for p in top_10_critical]
    chart_forecast = [p['forecast_30_days'] for p in top_10_critical]
    
    context = {
        'top_10_critical': top_10_critical,
        'page_obj': page_obj,
        'chart_labels': json.dumps(chart_labels),
        'chart_sold': json.dumps(chart_sold),
        'chart_forecast': json.dumps(chart_forecast),
    }
    
    return render(request, 'inventory/stock_forecast.html', context)
